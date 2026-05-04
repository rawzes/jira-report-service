import logging
from typing import List, Optional, Dict
from datetime import datetime
from io import BytesIO
from collections import defaultdict
import tempfile
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from app.models.schemas import (
    ReportData, EmployeeMetrics, ProjectMetrics,
    IssueData, WorklogEntry, IssueTransition,
    WeeklyThroughput, WeeklyCycleTime, CycleTimeData, AgingWipData,
    OverdueData, ReopenedData, RiskFlag, StatusDistribution
)
from app.models.config import settings

logger = logging.getLogger(__name__)

# Localization strings
LOCALIZATION = {
    "en": {
        "summary": "Summary",
        "by_employee": "By Employee",
        "by_project": "By Project",
        "weekly_throughput": "Weekly Throughput",
        "cycle_time": "Cycle Time",
        "aging_wip": "Aging WIP",
        "overdue": "Overdue",
        "reopened": "Reopened",
        "raw_issues": "Raw Issues",
        "raw_worklogs": "Raw Worklogs",
        "raw_transitions": "Raw Transitions",
        "risks": "Risks",
        "metric": "Metric",
        "value": "Value",
        "report_period": "Report Period",
        "total_worklog_hours": "Total Worklog Hours",
        "total_created_issues": "Total Created Issues",
        "total_closed_issues": "Total Closed Issues",
        "throughput": "Throughput",
        "median_cycle_time": "Median Cycle Time (days)",
        "p85_cycle_time": "85th Percentile Cycle Time (days)",
        "overdue_tasks": "Overdue Tasks",
        "reopened_tasks": "Reopened Tasks",
        "blocked_tasks": "Blocked Tasks",
        "top_oldest_issues": "Top 5 Oldest Open Issues",
        "issue_key": "Issue Key",
        "summary": "Summary",
        "project_key": "Project Key",
        "status": "Status",
        "created": "Created",
        "aging_days": "Aging (days)",
        "assignee": "Assignee",
        "employee_name": "Employee Name",
        "account_id": "Account ID",
        "worklog_hours": "Worklog Hours",
        "created_issues": "Created Issues",
        "closed_issues": "Closed Issues",
        "reopen_count": "Reopened Count",
        "active_wip": "Active WIP",
        "avg_cycle_time": "Avg Cycle Time",
        "median_cycle_time": "Median Cycle Time",
        "overdue_count": "Overdue Count",
        "project": "Project",
        "throughput_count": "Throughput",
        "wip_count": "WIP Count",
        "aging_7": "Aging >7 days",
        "aging_14": "Aging >14 days",
        "aging_30": "Aging >30 days",
        "blocked_count": "Blocked Count",
        "week_start": "Week Start",
        "week_end": "Week End",
        "created_count": "Created Count",
        "closed_count": "Closed Count",
        "net_flow": "Net Flow",
        "issue_type": "Issue Type",
        "reporter": "Reporter",
        "start_progress_date": "Start Progress Date",
        "done_date": "Done Date",
        "cycle_time_days": "Cycle Time (days)",
        "cycle_time_hours": "Cycle Time (hours)",
        "reopened_flag": "Reopened?",
        "last_status_change": "Last Status Change",
        "blocked_flag": "Blocked?",
        "due_date": "Due Date",
        "overdue_flag": "Overdue?",
        "priority": "Priority",
        "days_overdue": "Days Overdue",
        "reopen_date": "Reopen Date",
        "current_status": "Current Status",
        "from_status": "From Status",
        "to_status": "To Status",
        "transition_date": "Transition Date",
        "transition_author": "Transition Author",
        "started": "Started",
        "time_spent_hours": "Time Spent (hours)",
        "comment": "Comment",
        "risk_type": "Risk Type",
        "description": "Description",
        "severity": "Severity",
        "total": "TOTAL",
        "cycle_time_by_week": "Cycle Time by Week",
        "status_by_project": "Status by Project",
        "charts": "Charts",
        "closed_by_week": "Closed Tasks by Week",
        "cycle_time_by_week": "Cycle Time by Week",
        "worklog_by_employee": "Worklog by Employee",
        "aging_buckets": "Aging WIP Buckets",
        "status_by_project": "Status by Project",
        "reopened_by_project": "Reopened by Project",
    },
    "ru": {
        "summary": "Сводка",
        "by_employee": "По сотрудникам",
        "by_project": "По проектам",
        "weekly_throughput": "Недельный Throughput",
        "cycle_time": "Cycle Time",
        "aging_wip": "Aging WIP",
        "overdue": "Просроченные",
        "reopened": "Переоткрытые",
        "raw_issues": "Сырые данные задач",
        "raw_worklogs": "Сырые данные трудозатрат",
        "raw_transitions": "Сырые данные переходов",
        "risks": "Риски",
        "metric": "Показатель",
        "value": "Значение",
        "report_period": "Период отчета",
        "total_worklog_hours": "Всего часов трудозатрат",
        "total_created_issues": "Всего создано задач",
        "total_closed_issues": "Всего закрыто задач",
        "throughput": "Throughput",
        "median_cycle_time": "Медиана Cycle Time (дни)",
        "p85_cycle_time": "85-й перцентиль Cycle Time (дни)",
        "overdue_tasks": "Просроченные задачи",
        "reopened_tasks": "Переоткрытые задачи",
        "blocked_tasks": "Заблокированные задачи",
        "top_oldest_issues": "Топ-5 самых старых открытых задач",
        "issue_key": "Ключ задачи",
        "summary": "Название",
        "project_key": "Проект",
        "status": "Статус",
        "created": "Создана",
        "aging_days": "Возраст (дни)",
        "assignee": "Исполнитель",
        "employee_name": "Сотрудник",
        "account_id": "ID аккаунта",
        "worklog_hours": "Часы трудозатрат",
        "created_issues": "Создано задач",
        "closed_issues": "Закрыто задач",
        "reopen_count": "Кол-во переоткрытий",
        "active_wip": "Активный WIP",
        "avg_cycle_time": "Средний Cycle Time",
        "median_cycle_time": "Медиана Cycle Time",
        "overdue_count": "Просрочено",
        "project": "Проект",
        "throughput_count": "Throughput",
        "wip_count": "WIP",
        "aging_7": "Возраст >7 дней",
        "aging_14": "Возраст >14 дней",
        "aging_30": "Возраст >30 дней",
        "blocked_count": "Заблокировано",
        "week_start": "Начало недели",
        "week_end": "Конец недели",
        "created_count": "Создано",
        "closed_count": "Закрыто",
        "net_flow": "Чистый поток",
        "issue_type": "Тип задачи",
        "reporter": "Репортер",
        "start_progress_date": "Дата начала работы",
        "done_date": "Дата завершения",
        "cycle_time_days": "Cycle Time (дни)",
        "cycle_time_hours": "Cycle Time (часы)",
        "reopened_flag": "Переоткрыта?",
        "last_status_change": "Последний переход",
        "blocked_flag": "Заблокирована?",
        "due_date": "Срок",
        "overdue_flag": "Просрочена?",
        "priority": "Приоритет",
        "days_overdue": "Дней просрочки",
        "reopen_date": "Дата переоткрытия",
        "current_status": "Текущий статус",
        "from_status": "Из статуса",
        "to_status": "В статус",
        "transition_date": "Дата перехода",
        "transition_author": "Автор перехода",
        "started": "Начало",
        "time_spent_hours": "Часы",
        "comment": "Комментарий",
        "risk_type": "Тип риска",
        "description": "Описание",
        "severity": "Серьезность",
        "total": "ИТОГО",
        "cycle_time_by_week": "Cycle Time по неделям",
        "status_by_project": "Статусы по проектам",
        "charts": "Графики",
        "closed_by_week": "Закрытые задачи по неделям",
        "cycle_time_by_week": "Cycle Time по неделям",
        "worklog_by_employee": "Трудозатраты по сотрудникам",
        "aging_buckets": "Возрастные группы WIP",
        "status_by_project": "Статусы по проектам",
        "reopened_by_project": "Переоткрытые по проектам",
    }
}


class XlsxExporter:
    """Generates Excel reports with multiple sheets, charts, and formatting."""
    
    def __init__(self, lang: str = None):
        self.lang = lang or settings.REPORT_LANG
        if self.lang not in LOCALIZATION:
            self.lang = "en"
        self._ = LOCALIZATION[self.lang]
        
        # Pre-create style objects for reuse
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.total_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.total_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        self.alt_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.number_format = '0.00'
        self.date_format = "yyyy-mm-dd"
        
    def _naive_datetime(self, dt):
        """Convert timezone-aware datetime to naive datetime for Excel compatibility."""
        if dt is None:
            return None
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt
        
    def generate(self, report: ReportData) -> BytesIO:
        """Generate XLSX report with all sheets and charts."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_summary_sheet(wb, report)
        self._create_by_employee_sheet(wb, report)
        self._create_by_project_sheet(wb, report)
        self._create_weekly_throughput_sheet(wb, report)
        self._create_cycle_time_sheet(wb, report)
        self._create_aging_wip_sheet(wb, report)
        self._create_overdue_sheet(wb, report)
        self._create_reopened_sheet(wb, report)
        
        # Raw data sheets (if included)
        if report.include_raw_data:
            self._create_raw_issues_sheet(wb, report)
            self._create_raw_worklogs_sheet(wb, report)
            self._create_raw_transitions_sheet(wb, report)
        
        # Risks sheet
        self._create_risks_sheet(wb, report)
        
        # Charts sheet
        self._create_charts_sheet(wb, report)
        
        # Save to BytesIO

        # Convert all timezone-aware datetimes to naive (brute-force fix)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and hasattr(cell.value, "tzinfo") and cell.value.tzinfo is not None:
                        cell.value = cell.value.replace(tzinfo=None)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _format_header_row(self, ws, row_num=1, fill_color=None):
        """Format a header row."""
        fill = PatternFill(start_color=fill_color or "4472C4", end_color=fill_color or "4472C4", fill_type="solid")
        for cell in ws[row_num]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    def _format_data_row(self, ws, row_num, num_cols=None, alt=False):
        """Format a data row."""
        if num_cols is None:
            num_cols = ws.max_column
        fill = self.alt_fill if alt else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row_num, col)
            cell.border = self.border
            if fill:
                cell.fill = fill
            if isinstance(cell.value, (int, float)) and col > 1:
                cell.number_format = self.number_format
    
    def _format_integer_columns(self, ws, col_indices, min_row=2):
        """Set integer format for specified columns."""
        for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row):
            for col_idx in col_indices:
                if col_idx <= len(row):
                    cell = row[col_idx - 1]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
    
    def _set_column_widths(self, ws, widths):
        """Set column widths."""
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    
    def _create_summary_sheet(self, wb: Workbook, report: ReportData):
        """Create Summary sheet with key metrics."""
        ws = wb.create_sheet(self._["summary"])
        
        # Title
        ws['A1'] = self._["report_period"]
        ws['B1'] = f"{report.start_date.strftime('%Y-%m-%d')} - {report.end_date.strftime('%Y-%m-%d')}"
        ws.merge_cells('B1:D1')
        
        # Headers
        ws['A3'] = self._["metric"]
        ws['B3'] = self._["value"]
        self._format_header_row(ws, 3)
        
        # Metrics
        metrics = [
            (self._["total_worklog_hours"], report.total_worklog_hours),
            (self._["total_created_issues"], report.total_created_issues),
            (self._["total_closed_issues"], report.total_closed_issues),
            (self._["throughput"], report.throughput),
            (self._["median_cycle_time"], report.median_cycle_time),
            (self._["p85_cycle_time"], report.p85_cycle_time),
            (self._["overdue_tasks"], report.overdue_tasks_count),
            (self._["reopened_tasks"], report.reopened_tasks_count),
            (self._["blocked_tasks"], report.blocked_tasks_count),
        ]
        
        for idx, (metric, value) in enumerate(metrics, 4):
            ws[f'A{idx}'] = metric
            ws[f'B{idx}'] = value
            self._format_data_row(ws, idx, 2, alt=(idx % 2 == 0))
        
        # Top 5 oldest issues
        if report.top_oldest_issues:
            start_row = len(metrics) + 6
            ws[f'A{start_row}'] = self._["top_oldest_issues"]
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
            
            # Headers
            headers = [self._["issue_key"], self._["summary"], self._["project_key"], 
                      self._["status"], self._["aging_days"], self._["assignee"]]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row+1, column=col)
                cell.value = header
            self._format_header_row(ws, start_row+1)
            
            for idx, issue in enumerate(report.top_oldest_issues, start_row+2):
                ws.cell(row=idx, column=1, value=issue.get("issue_key", ""))
                ws.cell(row=idx, column=2, value=issue.get("summary", ""))
                ws.cell(row=idx, column=3, value=issue.get("project_key", ""))
                ws.cell(row=idx, column=4, value=issue.get("status", ""))
                ws.cell(row=idx, column=5, value=issue.get("aging_days", 0))
                ws.cell(row=idx, column=6, value=issue.get("assignee", ""))
                self._format_data_row(ws, idx, 6, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [40, 30])
    
    def _create_by_employee_sheet(self, wb: Workbook, report: ReportData):
        """Create By Employee sheet."""
        ws = wb.create_sheet(self._["by_employee"])
        
        # Headers
        headers = [
            self._["employee_name"], self._["account_id"], self._["project_key"],
            self._["worklog_hours"], self._["created_issues"], self._["closed_issues"],
            self._["reopen_count"], self._["active_wip"], self._["avg_cycle_time"],
            self._["median_cycle_time"], self._["overdue_count"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, emp in enumerate(report.employee_metrics, 2):
            ws.cell(row=idx, column=1, value=emp.display_name)
            ws.cell(row=idx, column=2, value=emp.account_id)
            ws.cell(row=idx, column=3, value=emp.project_key)
            ws.cell(row=idx, column=4, value=emp.worklog_hours)
            ws.cell(row=idx, column=5, value=emp.created_issues)
            ws.cell(row=idx, column=6, value=emp.closed_issues)
            ws.cell(row=idx, column=7, value=emp.reopened_tasks)
            ws.cell(row=idx, column=8, value=emp.active_wip)
            ws.cell(row=idx, column=9, value=emp.avg_cycle_time)
            ws.cell(row=idx, column=10, value=emp.median_cycle_time)
            ws.cell(row=idx, column=11, value=emp.overdue_tasks)
            self._format_data_row(ws, idx, 11, alt=(idx % 2 == 0))
        
        # Total row
        total_row = len(report.employee_metrics) + 2
        ws.cell(row=total_row, column=1, value=self._["total"])
        ws.cell(row=total_row, column=4, value=sum(e.worklog_hours for e in report.employee_metrics))
        ws.cell(row=total_row, column=5, value=sum(e.created_issues for e in report.employee_metrics))
        ws.cell(row=total_row, column=6, value=sum(e.closed_issues for e in report.employee_metrics))
        ws.cell(row=total_row, column=7, value=sum(e.reopened_tasks for e in report.employee_metrics))
        ws.cell(row=total_row, column=8, value=sum(e.active_wip for e in report.employee_metrics))
        ws.cell(row=total_row, column=11, value=sum(e.overdue_tasks for e in report.employee_metrics))
        self._format_data_row(ws, total_row, 11)
        for col in range(1, 12):
            ws.cell(row=total_row, column=col).font = self.total_font
            ws.cell(row=total_row, column=col).fill = self.total_fill
        
        # Set column widths
        self._set_column_widths(ws, [25, 20, 12, 15, 15, 15, 15, 12, 18, 18, 12])
    
    def _create_by_project_sheet(self, wb: Workbook, report: ReportData):
        """Create By Project sheet."""
        ws = wb.create_sheet(self._["by_project"])
        
        # Headers
        headers = [
            self._["project_key"], self._["worklog_hours"], self._["created_issues"],
            self._["closed_issues"], self._["throughput_count"], self._["median_cycle_time"],
            self._["p85_cycle_time"], self._["wip_count"], self._["aging_7"],
            self._["aging_14"], self._["aging_30"], self._["reopen_count"],
            self._["overdue_count"], self._["blocked_count"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, proj in enumerate(report.project_metrics, 2):
            ws.cell(row=idx, column=1, value=proj.project_key)
            ws.cell(row=idx, column=2, value=proj.worklog_hours)
            ws.cell(row=idx, column=3, value=proj.created_issues)
            ws.cell(row=idx, column=4, value=proj.closed_issues)
            ws.cell(row=idx, column=5, value=proj.throughput)
            ws.cell(row=idx, column=6, value=proj.median_cycle_time)
            ws.cell(row=idx, column=7, value=proj.p85_cycle_time)
            ws.cell(row=idx, column=8, value=proj.wip_count)
            ws.cell(row=idx, column=9, value=proj.aging_wip_7)
            ws.cell(row=idx, column=10, value=proj.aging_wip_14)
            ws.cell(row=idx, column=11, value=proj.aging_wip_30)
            ws.cell(row=idx, column=12, value=proj.reopened_tasks)
            ws.cell(row=idx, column=13, value=proj.overdue_tasks)
            ws.cell(row=idx, column=14, value=proj.blocked_tasks)
            self._format_data_row(ws, idx, 14, alt=(idx % 2 == 0))
        
        # Total row
        total_row = len(report.project_metrics) + 2
        ws.cell(row=total_row, column=1, value=self._["total"])
        ws.cell(row=total_row, column=2, value=sum(p.worklog_hours for p in report.project_metrics))
        ws.cell(row=total_row, column=3, value=sum(p.created_issues for p in report.project_metrics))
        ws.cell(row=total_row, column=4, value=sum(p.closed_issues for p in report.project_metrics))
        ws.cell(row=total_row, column=5, value=sum(p.throughput for p in report.project_metrics))
        ws.cell(row=total_row, column=8, value=sum(p.wip_count for p in report.project_metrics))
        ws.cell(row=total_row, column=12, value=sum(p.reopened_tasks for p in report.project_metrics))
        ws.cell(row=total_row, column=13, value=sum(p.overdue_tasks for p in report.project_metrics))
        ws.cell(row=total_row, column=14, value=sum(p.blocked_tasks for p in report.project_metrics))
        self._format_data_row(ws, total_row, 14)
        for col in range(1, 15):
            ws.cell(row=total_row, column=col).font = self.total_font
            ws.cell(row=total_row, column=col).fill = self.total_fill
        
        # Set column widths
        self._set_column_widths(ws, [12, 15, 15, 15, 15, 18, 18, 12, 12, 12, 12, 15, 15, 15])
    
    def _create_weekly_throughput_sheet(self, wb: Workbook, report: ReportData):
        """Create Weekly Throughput sheet."""
        ws = wb.create_sheet(self._["weekly_throughput"])
        
        # Headers
        headers = [
            self._["week_start"], self._["week_end"], self._["project_key"],
            self._["closed_count"], self._["created_count"], self._["net_flow"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, wt in enumerate(report.weekly_throughput, 2):
            ws.cell(row=idx, column=1, value=self._naive_datetime(wt.week_start))
            ws.cell(row=idx, column=1).number_format = self.date_format
            ws.cell(row=idx, column=2, value=self._naive_datetime(wt.week_end))
            ws.cell(row=idx, column=2).number_format = self.date_format
            ws.cell(row=idx, column=3, value=wt.project_key)
            ws.cell(row=idx, column=4, value=wt.closed_tasks_count)
            ws.cell(row=idx, column=5, value=wt.created_tasks_count)
            ws.cell(row=idx, column=6, value=wt.net_flow)
            self._format_data_row(ws, idx, 6, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 15, 12, 15, 15, 12])
    
    def _create_cycle_time_sheet(self, wb: Workbook, report: ReportData):
        """Create Cycle Time sheet."""
        ws = wb.create_sheet(self._["cycle_time"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["issue_type"],
            self._["assignee"], self._["reporter"], self._["start_progress_date"],
            self._["done_date"], self._["cycle_time_days"], self._["cycle_time_hours"],
            self._["reopened_flag"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, ct in enumerate(report.cycle_time_data, 2):
            ws.cell(row=idx, column=1, value=ct.issue_key)
            ws.cell(row=idx, column=2, value=ct.project_key)
            ws.cell(row=idx, column=3, value=ct.issue_type)
            ws.cell(row=idx, column=4, value=ct.assignee)
            ws.cell(row=idx, column=5, value=ct.reporter)
            ws.cell(row=idx, column=6, value=self._naive_datetime(ct.start_progress_date))
            if ct.start_progress_date:
                ws.cell(row=idx, column=6).number_format = self.date_format
            ws.cell(row=idx, column=7, value=self._naive_datetime(ct.done_date))
            if ct.done_date:
                ws.cell(row=idx, column=7).number_format = self.date_format
            ws.cell(row=idx, column=8, value=ct.cycle_time_days)
            ws.cell(row=idx, column=9, value=ct.cycle_time_hours)
            ws.cell(row=idx, column=10, value="Yes" if ct.reopened_flag else "No")
            self._format_data_row(ws, idx, 10, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 15, 20, 20, 18, 18, 18, 18, 12])
    
    def _create_aging_wip_sheet(self, wb: Workbook, report: ReportData):
        """Create Aging WIP sheet."""
        ws = wb.create_sheet(self._["aging_wip"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["status"],
            self._["assignee"], self._["created"], self._["last_status_change"],
            self._["aging_days"], self._["blocked_flag"], self._["due_date"],
            self._["overdue_flag"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, aw in enumerate(report.aging_wip_data, 2):
            ws.cell(row=idx, column=1, value=aw.issue_key)
            ws.cell(row=idx, column=2, value=aw.project_key)
            ws.cell(row=idx, column=3, value=aw.status)
            ws.cell(row=idx, column=4, value=aw.assignee)
            ws.cell(row=idx, column=5, value=self._naive_datetime(aw.created_date))
            if aw.created_date:
                ws.cell(row=idx, column=5).number_format = self.date_format
            ws.cell(row=idx, column=6, value=self._naive_datetime(aw.last_status_change_date))
            if aw.last_status_change_date:
                ws.cell(row=idx, column=6).number_format = self.date_format
            ws.cell(row=idx, column=7, value=aw.aging_days)
            ws.cell(row=idx, column=8, value="Yes" if aw.blocked_flag else "No")
            ws.cell(row=idx, column=9, value=self._naive_datetime(aw.due_date))
            if aw.due_date:
                ws.cell(row=idx, column=9).number_format = self.date_format
            ws.cell(row=idx, column=10, value="Yes" if aw.overdue_flag else "No")
            self._format_data_row(ws, idx, 10, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 15, 20, 18, 18, 12, 12, 18, 12])
    
    def _create_overdue_sheet(self, wb: Workbook, report: ReportData):
        """Create Overdue sheet."""
        ws = wb.create_sheet(self._["overdue"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["assignee"],
            self._["status"], self._["due_date"], self._["days_overdue"],
            self._["priority"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, od in enumerate(report.overdue_data, 2):
            ws.cell(row=idx, column=1, value=od.issue_key)
            ws.cell(row=idx, column=2, value=od.project_key)
            ws.cell(row=idx, column=3, value=od.assignee)
            ws.cell(row=idx, column=4, value=od.status)
            ws.cell(row=idx, column=5, value=self._naive_datetime(od.due_date))
            if od.due_date:
                ws.cell(row=idx, column=5).number_format = self.date_format
            ws.cell(row=idx, column=6, value=od.days_overdue)
            ws.cell(row=idx, column=7, value=od.priority)
            self._format_data_row(ws, idx, 7, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 20, 15, 18, 15, 15])
    
    def _create_reopened_sheet(self, wb: Workbook, report: ReportData):
        """Create Reopened sheet."""
        ws = wb.create_sheet(self._["reopened"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["assignee"],
            self._["done_date"], self._["reopen_date"], self._["reopen_count"],
            self._["current_status"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, rd in enumerate(report.reopened_data, 2):
            ws.cell(row=idx, column=1, value=rd.issue_key)
            ws.cell(row=idx, column=2, value=rd.project_key)
            ws.cell(row=idx, column=3, value=rd.assignee)
            ws.cell(row=idx, column=4, value=self._naive_datetime(rd.done_date))
            if rd.done_date:
                ws.cell(row=idx, column=4).number_format = self.date_format
            ws.cell(row=idx, column=5, value=self._naive_datetime(rd.reopen_date))
            if rd.reopen_date:
                ws.cell(row=idx, column=5).number_format = self.date_format
            ws.cell(row=idx, column=6, value=rd.reopen_count)
            ws.cell(row=idx, column=7, value=rd.current_status)
            self._format_data_row(ws, idx, 7, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 20, 18, 18, 12, 15])
    
    def _create_raw_issues_sheet(self, wb: Workbook, report: ReportData):
        """Create Raw Issues sheet."""
        ws = wb.create_sheet(self._["raw_issues"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["issue_type"],
            self._["summary"], self._["status"], self._["created"],
            self._["assignee"], self._["reporter"], self._["due_date"],
            self._["priority"], self._["cycle_time_days"], self._["reopened_flag"],
            self._["blocked_flag"], self._["overdue_flag"], self._["aging_days"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, issue in enumerate(report.raw_issues, 2):
            ws.cell(row=idx, column=1, value=issue.issue_key)
            ws.cell(row=idx, column=2, value=issue.project_key)
            ws.cell(row=idx, column=3, value=issue.issue_type)
            ws.cell(row=idx, column=4, value=issue.summary)
            ws.cell(row=idx, column=5, value=issue.status)
            ws.cell(row=idx, column=6, value=self._naive_datetime(issue.created))
            if issue.created:
                ws.cell(row=idx, column=6).number_format = self.date_format
            ws.cell(row=idx, column=7, value=issue.assignee_display_name)
            ws.cell(row=idx, column=8, value=issue.reporter)
            ws.cell(row=idx, column=9, value=self._naive_datetime(issue.due_date))
            if issue.due_date:
                ws.cell(row=idx, column=9).number_format = self.date_format
            ws.cell(row=idx, column=10, value=issue.priority)
            ws.cell(row=idx, column=11, value=issue.cycle_time_days)
            ws.cell(row=idx, column=12, value="Yes" if issue.reopened_flag else "No")
            ws.cell(row=idx, column=13, value="Yes" if issue.blocked_flag else "No")
            ws.cell(row=idx, column=14, value="Yes" if issue.overdue_flag else "No")
            ws.cell(row=idx, column=15, value=issue.aging_days)
            self._format_data_row(ws, idx, 15, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 15, 40, 15, 18, 20, 20, 18, 15, 18, 12, 12, 12, 12])
    
    def _create_raw_worklogs_sheet(self, wb: Workbook, report: ReportData):
        """Create Raw Worklogs sheet."""
        ws = wb.create_sheet(self._["raw_worklogs"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["employee_name"],
            self._["started"], self._["time_spent_hours"], self._["comment"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, wl in enumerate(report.raw_worklogs, 2):
            ws.cell(row=idx, column=1, value=wl.issue_key)
            ws.cell(row=idx, column=2, value=wl.project_key)
            ws.cell(row=idx, column=3, value=wl.display_name)
            ws.cell(row=idx, column=4, value=self._naive_datetime(wl.started))
            if wl.started:
                ws.cell(row=idx, column=4).number_format = self.date_format
            ws.cell(row=idx, column=5, value=wl.time_spent_hours)
            ws.cell(row=idx, column=6, value=wl.comment)
            self._format_data_row(ws, idx, 6, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 20, 18, 15, 40])
    
    def _create_raw_transitions_sheet(self, wb: Workbook, report: ReportData):
        """Create Raw Transitions sheet."""
        ws = wb.create_sheet(self._["raw_transitions"])
        
        # Headers
        headers = [
            self._["issue_key"], self._["project_key"], self._["from_status"],
            self._["to_status"], self._["transition_date"], self._["transition_author"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, trans in enumerate(report.raw_transitions, 2):
            ws.cell(row=idx, column=1, value=trans.issue_key)
            ws.cell(row=idx, column=2, value=trans.project_key)
            ws.cell(row=idx, column=3, value=trans.from_status)
            ws.cell(row=idx, column=4, value=trans.to_status)
            ws.cell(row=idx, column=5, value=self._naive_datetime(trans.transition_date))
            if trans.transition_date:
                ws.cell(row=idx, column=5).number_format = self.date_format
            ws.cell(row=idx, column=6, value=trans.transition_author)
            self._format_data_row(ws, idx, 6, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [15, 12, 15, 15, 18, 20])
    
    def _create_risks_sheet(self, wb: Workbook, report: ReportData):
        """Create Risks sheet."""
        ws = wb.create_sheet(self._["risks"])
        
        # Headers
        headers = [
            self._["risk_type"], self._["description"], self._["severity"],
            self._["project_key"], self._["employee_name"]
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._format_header_row(ws, 1)
        
        # Data rows
        for idx, risk in enumerate(report.risk_flags, 2):
            ws.cell(row=idx, column=1, value=risk.risk_type)
            ws.cell(row=idx, column=2, value=risk.description)
            ws.cell(row=idx, column=3, value=risk.severity)
            ws.cell(row=idx, column=4, value=risk.project_key)
            ws.cell(row=idx, column=5, value=risk.employee_name)
            self._format_data_row(ws, idx, 5, alt=(idx % 2 == 0))
        
        # Set column widths
        self._set_column_widths(ws, [20, 60, 12, 15, 20])
    
    def _create_charts_sheet(self, wb: Workbook, report: ReportData):
        """Create Charts sheet with embedded chart images."""
        ws = wb.create_sheet(self._["charts"])
        
        row = 1
        temp_files = []
        
        try:
            # 1. Closed tasks by week (Bar chart)
            ws[f'A{row}'] = self._["closed_by_week"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_closed_by_week_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 20
            
            # 2. Worklog hours by employee (Bar chart)
            ws[f'A{row}'] = self._["worklog_by_employee"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_worklog_by_employee_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 20
            
            # 3. Aging WIP buckets (Bar chart)
            ws[f'A{row}'] = self._["aging_buckets"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_aging_buckets_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 400
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 20
            
            # 4. Reopened tasks by project (Bar chart)
            ws[f'A{row}'] = self._["reopened_by_project"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_reopened_by_project_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 500
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 20
            
            # 5. Weekly Cycle Time (Line chart)
            ws[f'A{row}'] = self._["cycle_time_by_week"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_cycle_time_by_week_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 20
            
            # 6. Status Distribution by Project (Stacked bar chart)
            ws[f'A{row}'] = self._["status_by_project"]
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            img_path = self._generate_status_by_project_chart(report)
            if img_path:
                temp_files.append(img_path)
                img = XlImage(img_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'A{row}')
        
        finally:
            # Clean up temp files
            for f in temp_files:
                try:
                    os.unlink(f)
                except:
                    pass

    def _save_chart_to_temp(self, fig) -> Optional[str]:
        """Save matplotlib figure to a temporary file and return the path."""
        try:
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig.savefig(tmp.name, format='png', dpi=100, bbox_inches='tight')
                plt.close(fig)
                return tmp.name
        except Exception as e:
            logger.error(f"Failed to save chart to temp file: {e}")
            plt.close(fig)
            return None

    def _generate_closed_by_week_chart(self, report: ReportData) -> Optional[str]:
        """Generate closed tasks by week chart."""
        try:
            weekly_data = defaultdict(lambda: defaultdict(int))
            for wt in report.weekly_throughput:
                week_label = wt.week_start.strftime('%Y-%m-%d')
                weekly_data[wt.project_key][week_label] = wt.closed_tasks_count
            
            projects = sorted(set(wt.project_key for wt in report.weekly_throughput))
            weeks = sorted(set(wt.week_start.strftime('%Y-%m-%d') for wt in report.weekly_throughput))
            
            if not weeks or not projects:
                return None
            
            fig, ax = plt.subplots(figsize=(10, 5))
            
            x = range(len(weeks))
            width = 0.8 / len(projects) if projects else 0.8
            
            for i, proj in enumerate(projects):
                values = [weekly_data[proj].get(week, 0) for week in weeks]
                ax.bar([pos + i * width for pos in x], values, width, label=proj)
            
            ax.set_xlabel('Week')
            ax.set_ylabel('Closed Tasks')
            ax.set_title(self._["closed_by_week"])
            ax.set_xticks([pos + width * len(projects) / 2 for pos in x])
            ax.set_xticklabels(weeks, rotation=45, ha='right')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate closed by week chart: {e}")
            return None

    def _generate_worklog_by_employee_chart(self, report: ReportData) -> Optional[str]:
        """Generate worklog hours by employee chart."""
        try:
            if not report.employee_metrics:
                return None
            
            employees = [emp.display_name for emp in report.employee_metrics]
            hours = [emp.worklog_hours for emp in report.employee_metrics]
            
            fig, ax = plt.subplots(figsize=(10, 6))
            bars = ax.bar(range(len(employees)), hours)
            ax.set_xlabel('Employee')
            ax.set_ylabel('Hours')
            ax.set_title(self._["worklog_by_employee"])
            ax.set_xticks(range(len(employees)))
            ax.set_xticklabels(employees, rotation=45, ha='right')
            
            # Add value labels on bars
            for bar, hour in zip(bars, hours):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{hour:.1f}', ha='center', va='bottom')
            
            ax.grid(True, alpha=0.3, axis='y')
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate worklog by employee chart: {e}")
            return None

    def _generate_aging_buckets_chart(self, report: ReportData) -> Optional[str]:
        """Generate aging WIP buckets chart."""
        try:
            buckets = {"0-7": 0, "8-14": 0, "15-30": 0, "30+": 0}
            for aw in report.aging_wip_data:
                if aw.aging_days <= 7:
                    buckets["0-7"] += 1
                elif aw.aging_days <= 14:
                    buckets["8-14"] += 1
                elif aw.aging_days <= 30:
                    buckets["15-30"] += 1
                else:
                    buckets["30+"] += 1
            
            fig, ax = plt.subplots(figsize=(8, 5))
            bars = ax.bar(buckets.keys(), buckets.values())
            ax.set_xlabel('Days')
            ax.set_ylabel('Count')
            ax.set_title(self._["aging_buckets"])
            
            # Add value labels on bars
            for bar, count in zip(bars, buckets.values()):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       str(count), ha='center', va='bottom')
            
            ax.grid(True, alpha=0.3, axis='y')
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate aging buckets chart: {e}")
            return None

    def _generate_reopened_by_project_chart(self, report: ReportData) -> Optional[str]:
        """Generate reopened tasks by project chart."""
        try:
            if not report.project_metrics:
                return None
            
            projects = [proj.project_key for proj in report.project_metrics]
            reopened = [proj.reopened_tasks for proj in report.project_metrics]
            
            fig, ax = plt.subplots(figsize=(8, 5))
            bars = ax.bar(projects, reopened)
            ax.set_xlabel('Project')
            ax.set_ylabel('Reopened Count')
            ax.set_title(self._["reopened_by_project"])
            
            # Add value labels on bars
            for bar, count in zip(bars, reopened):
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           str(count), ha='center', va='bottom')
            
            ax.grid(True, alpha=0.3, axis='y')
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate reopened by project chart: {e}")
            return None

    def _generate_cycle_time_by_week_chart(self, report: ReportData) -> Optional[str]:
        """Generate cycle time by week chart."""
        try:
            week_labels = sorted(set(
                wt.week_start.strftime('%Y-%m-%d')
                for wt in report.weekly_cycle_time
            ))
            projects = sorted(set(wt.project_key for wt in report.weekly_cycle_time))
            
            if not week_labels or not projects:
                return None
            
            fig, ax = plt.subplots(figsize=(10, 5))
            
            for proj in projects:
                values = []
                for week in week_labels:
                    ct = next(
                        (wt for wt in report.weekly_cycle_time
                        if wt.week_start.strftime('%Y-%m-%d') == week and wt.project_key == proj
                    ), None)
                    values.append(ct.median_cycle_time if ct else None)
                
                # Filter out None values for plotting
                x_vals = [week_labels[i] for i, v in enumerate(values) if v is not None]
                y_vals = [v for v in values if v is not None]
                
                if y_vals:
                    ax.plot(x_vals, y_vals, marker='o', label=proj)
            
            ax.set_xlabel('Week')
            ax.set_ylabel('Days')
            ax.set_title(self._["cycle_time_by_week"])
            plt.xticks(rotation=45, ha='right')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate cycle time by week chart: {e}")
            return None

    def _generate_status_by_project_chart(self, report: ReportData) -> Optional[str]:
        """Generate status distribution by project chart."""
        try:
            all_statuses = sorted(set(sd.status for sd in report.status_distribution))
            all_projects = sorted(set(sd.project_key for sd in report.status_distribution))
            
            if not all_statuses or not all_projects:
                return None
            
            # Prepare data
            data = defaultdict(lambda: defaultdict(int))
            for sd in report.status_distribution:
                data[sd.project_key][sd.status] = sd.count
            
            fig, ax = plt.subplots(figsize=(10, 6))
            
            bottom = [0] * len(all_projects)
            colors = plt.cm.Set3(range(len(all_statuses)))
            
            for i, status in enumerate(all_statuses):
                values = [data[proj].get(status, 0) for proj in all_projects]
                ax.bar(all_projects, values, bottom=bottom, label=status, color=colors[i])
                bottom = [b + v for b, v in zip(bottom, values)]
            
            ax.set_xlabel('Project')
            ax.set_ylabel('Count')
            ax.set_title(self._["status_by_project"])
            plt.xticks(rotation=45, ha='right')
            ax.legend()
            ax.grid(True, alpha=0.3, axis='y')
            
            return self._save_chart_to_temp(fig)
        except Exception as e:
            logger.error(f"Failed to generate status by project chart: {e}")
            return None
